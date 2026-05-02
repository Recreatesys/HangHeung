import base64
import io
import logging
import re

from odoo import models, fields, _, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

ALLOWED_PREFIXES = ('HHC', 'BWC', 'CWC', 'EWC', 'DPC', 'SWC', 'LWC')
TRAILING_DIGITS = 2
BATCH_SIZE = 2000


class SecurityCodeImportWizard(models.TransientModel):
    _name = 'loyalty.security.code.import.wizard'
    _description = 'Import Coupon Security Codes from Excel'

    file_data = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')

    # Pre-parsed rows kept across the JS-orchestrated batch calls.
    parsed_rows = fields.Json(string='Parsed Rows', readonly=True)
    total_rows = fields.Integer(string='Total Rows', readonly=True)
    total_batches = fields.Integer(string='Total Batches', readonly=True)

    # Running counters updated batch-by-batch.
    processed_count = fields.Integer(readonly=True)
    updated_count = fields.Integer(readonly=True)
    no_change_count = fields.Integer(readonly=True)
    skipped_count = fields.Integer(readonly=True)
    errors_log = fields.Text(readonly=True)

    result_summary = fields.Text(string='Import Result', readonly=True)

    def action_parse_file(self):
        """Read the uploaded Excel and stage rows for batched processing.

        Returns dict with total_rows and total_batches so the JS orchestrator
        knows how many process_batch calls to make.
        """
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Please upload an Excel file first."))

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("openpyxl is required to import Excel files."))

        try:
            wb = load_workbook(
                io.BytesIO(base64.b64decode(self.file_data)),
                read_only=True,
                data_only=True,
            )
        except Exception as e:
            raise UserError(_("Failed to open Excel file: %s") % e)

        sheet = wb.active
        all_rows = list(sheet.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            raise UserError(_("Excel file is empty."))

        header_skipped = False
        if all_rows and all_rows[0]:
            first = str(all_rows[0][0]).strip() if all_rows[0][0] is not None else ''
            if first and not re.fullmatch(r'[A-Z]{3}\d+', first):
                all_rows = all_rows[1:]
                header_skipped = True

        rows = []
        start_row_no = 2 if header_skipped else 1
        for offset, row in enumerate(all_rows):
            row_no = start_row_no + offset
            if not row or all(c is None or c == '' for c in row[:2]):
                continue
            code_val = str(row[0]).strip() if row[0] is not None else ''
            sec_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            rows.append({'row_no': row_no, 'code': code_val, 'sec': sec_val})

        total = len(rows)
        batches = (total + BATCH_SIZE - 1) // BATCH_SIZE if total else 0

        self.write({
            'parsed_rows': rows,
            'total_rows': total,
            'total_batches': batches,
            'processed_count': 0,
            'updated_count': 0,
            'no_change_count': 0,
            'skipped_count': 0,
            'errors_log': '',
            'result_summary': '',
        })

        _logger.info(
            "Security code import: parsed %d rows in %d batches (size=%d) for wizard %s",
            total, batches, BATCH_SIZE, self.id,
        )

        return {'total_rows': total, 'total_batches': batches}

    def action_process_batch(self, batch_idx):
        """Process a single 0-indexed batch of BATCH_SIZE rows."""
        self.ensure_one()
        rows = self.parsed_rows or []
        if not rows:
            raise UserError(_("No parsed rows. Call action_parse_file first."))

        start = batch_idx * BATCH_SIZE
        end = min(start + BATCH_SIZE, len(rows))
        if start >= len(rows):
            return {'batch_idx': batch_idx, 'processed_in_batch': 0}

        chunk = rows[start:end]

        Card = self.env['loyalty.card']
        upd = nochange = skip = 0
        errs = []

        for r in chunk:
            row_no = r.get('row_no')
            code_val = r.get('code') or ''
            sec_val = r.get('sec') or ''

            if not code_val:
                skip += 1
                errs.append(_("Row %s: empty coupon code") % row_no)
                continue
            if not sec_val:
                skip += 1
                errs.append(_("Row %s (%s): empty security code") % (row_no, code_val))
                continue

            prefix = code_val[:3].upper()
            if prefix not in ALLOWED_PREFIXES:
                skip += 1
                errs.append(_(
                    "Row %s (%s): prefix '%s' not in allowed set %s"
                ) % (row_no, code_val, prefix, ', '.join(ALLOWED_PREFIXES)))
                continue

            expected_pattern = re.escape(code_val) + r'\d{%d}' % TRAILING_DIGITS
            if not re.fullmatch(expected_pattern, sec_val):
                skip += 1
                errs.append(_(
                    "Row %s (%s): security code '%s' must be coupon code + %s digits"
                ) % (row_no, code_val, sec_val, TRAILING_DIGITS))
                continue

            card = Card.search([('code', '=', code_val)], limit=1)
            if not card:
                skip += 1
                errs.append(_("Row %s (%s): coupon not found in DB") % (row_no, code_val))
                continue

            if card.security_code == sec_val:
                nochange += 1
                continue

            try:
                card.write({'security_code': sec_val})
                upd += 1
            except Exception as e:
                skip += 1
                errs.append(_("Row %s (%s): %s") % (row_no, code_val, str(e)))

        new_errors = (self.errors_log or '') + ('\n'.join(errs) + '\n' if errs else '')
        self.write({
            'processed_count': self.processed_count + len(chunk),
            'updated_count': self.updated_count + upd,
            'no_change_count': self.no_change_count + nochange,
            'skipped_count': self.skipped_count + skip,
            'errors_log': new_errors,
        })

        _logger.info(
            "Security code import: batch %d/%d done (processed=%d updated=%d nochange=%d skipped=%d)",
            batch_idx + 1, self.total_batches,
            self.processed_count, self.updated_count, self.no_change_count, self.skipped_count,
        )

        return {
            'batch_idx': batch_idx + 1,
            'processed_in_batch': len(chunk),
            'processed_total': self.processed_count,
            'updated': self.updated_count,
            'no_change': self.no_change_count,
            'skipped': self.skipped_count,
        }

    def action_finalize(self):
        """Compose result_summary, free parsed_rows, return final stats."""
        self.ensure_one()
        errors = (self.errors_log or '').rstrip('\n').split('\n') if self.errors_log else []
        errors = [e for e in errors if e]
        if len(errors) > 50:
            errors = errors[:50] + [_("... %s more errors omitted") % (len(errors) - 50)]

        summary_lines = [
            _("Total rows: %s") % self.total_rows,
            _("Updated: %s") % self.updated_count,
            _("Already up-to-date: %s") % self.no_change_count,
            _("Skipped: %s") % self.skipped_count,
        ]
        if errors:
            summary_lines.append('')
            summary_lines.append(_("Errors:"))
            summary_lines.extend(errors)

        self.write({
            'result_summary': '\n'.join(summary_lines),
            'parsed_rows': False,  # release the JSON payload
        })
        return {
            'updated': self.updated_count,
            'no_change': self.no_change_count,
            'skipped': self.skipped_count,
        }
